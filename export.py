import csv
from typing import List, Tuple
from models import Resident
import lang

_ADDRESS_CACHE: dict = {}


def _headers() -> list:
    return [
        lang.get("export_col_last"),
        lang.get("export_col_first"),
        lang.get("export_col_address"),
        lang.get("export_col_status"),
        lang.get("export_col_dob"),
        lang.get("export_col_baptism"),
        lang.get("export_col_marriage"),
        lang.get("export_col_death"),
        lang.get("export_col_notes"),
    ]


def _get_address_street(address_id: int) -> str:
    if address_id not in _ADDRESS_CACHE:
        import database as db
        for a in db.get_addresses():
            _ADDRESS_CACHE[a.id] = a.street
    return _ADDRESS_CACHE.get(address_id, "")


def _resident_row(r: Resident) -> list:
    status = lang.get("status_deceased") if r.status == "deceased" \
             else lang.get("status_active")
    return [
        r.last_name,
        r.first_name,
        _get_address_street(r.address_id),
        status,
        r.birth_date    or "",
        r.baptism_date  or "",
        r.marriage_date or "",
        r.death_date    or "",
        r.notes,
    ]


def export_csv(path: str, residents: List[Resident]):
    _ADDRESS_CACHE.clear()
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(_headers())
        for r in sorted(residents, key=lambda x: (x.last_name, x.first_name)):
            writer.writerow(_resident_row(r))


def export_excel(path: str, residents: List[Resident]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed.\n"
            "Run: pip install openpyxl\n"
            "Then try again."
        )

    _ADDRESS_CACHE.clear()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = lang.get("export_col_address") if lang.current() == "uk" else "Residents"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center")

    for col, header in enumerate(_headers(), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, r in enumerate(
        sorted(residents, key=lambda x: (x.last_name, x.first_name)), start=2
    ):
        for col_idx, value in enumerate(_resident_row(r), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(path)


# ── Import ────────────────────────────────────────────────────────────────────

# Accepted status values in any language → canonical DB value
_STATUS_MAP = {
    "active": "active", "deceased": "deceased",
    "активний": "active", "помер": "deceased",
}


def _cell(row, idx: int, default: str = "") -> str:
    return str(row[idx]).strip() if len(row) > idx else default


def _date_or_none(row, idx: int):
    v = _cell(row, idx)
    return v if v else None


def _import_rows(rows) -> Tuple[int, int]:
    """Insert rows from a parsed file (iterable of sequences). Returns (new, skipped)."""
    import database as db
    new_count = skip_count = 0
    for row in rows:
        last   = _cell(row, 0)
        first  = _cell(row, 1)
        street = _cell(row, 2)
        if not last or not first or not street:
            continue  # skip blank / header-like rows
        status = _STATUS_MAP.get(_cell(row, 3).lower(), "active")
        dob      = _date_or_none(row, 4)
        baptism  = _date_or_none(row, 5)
        marriage = _date_or_none(row, 6)
        death    = _date_or_none(row, 7)
        notes    = _cell(row, 8)

        addr_id = db.find_or_create_address(street)
        if db.resident_exists(addr_id, first, last):
            skip_count += 1
        else:
            db.add_resident(Resident(
                id=None, address_id=addr_id,
                first_name=first, last_name=last,
                birth_date=dob, baptism_date=baptism,
                marriage_date=marriage, death_date=death,
                status=status, notes=notes,
            ))
            new_count += 1
    return new_count, skip_count


def import_csv(path: str) -> Tuple[int, int]:
    """Import residents from a CSV file. Returns (new, skipped)."""
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header row
        rows = list(reader)
    return _import_rows(rows)


def import_excel(path: str) -> Tuple[int, int]:
    """Import residents from an Excel (.xlsx) file. Returns (new, skipped)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed.\n"
            "Run: pip install openpyxl\n"
            "Then try again."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=2, values_only=True)  # skip header row
    rows = [[str(c) if c is not None else "" for c in row] for row in rows_iter]
    wb.close()
    return _import_rows(rows)
