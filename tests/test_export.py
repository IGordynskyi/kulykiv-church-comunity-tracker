"""Tests for export.py — CSV/Excel export and import."""
import csv
import io
import os
import pytest
from unittest.mock import patch, MagicMock
from models import Resident, Address
import lang


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_resident(**kwargs):
    defaults = dict(
        id=1, address_id=1,
        first_name="Ivan", last_name="Kovalenko",
        birth_date=None, baptism_date=None,
        marriage_date=None, death_date=None,
        status="active", notes=""
    )
    defaults.update(kwargs)
    return Resident(**defaults)


@pytest.fixture(autouse=True)
def reset_lang_to_english():
    lang.set_lang("en")
    yield
    lang.set_lang("en")


@pytest.fixture
def db(tmp_path):
    """Isolated DB for tests that need import."""
    db_file = str(tmp_path / "test.db")
    with patch("database.DB_PATH", db_file):
        import database as _db
        _db.init_db()
        yield _db


# ── _cell / _date_or_none helpers ─────────────────────────────────────────────

class TestCellHelper:
    def test_returns_value_at_index(self):
        import export as exp
        assert exp._cell(["a", "b", "c"], 1) == "b"

    def test_strips_whitespace(self):
        import export as exp
        assert exp._cell(["  hello  ", "world"], 0) == "hello"

    def test_out_of_bounds_returns_default(self):
        import export as exp
        assert exp._cell(["a"], 5) == ""

    def test_out_of_bounds_custom_default(self):
        import export as exp
        assert exp._cell(["a"], 5, "N/A") == "N/A"

    def test_converts_non_string(self):
        import export as exp
        assert exp._cell([42, 99], 0) == "42"


class TestDateOrNone:
    def test_returns_value_when_present(self):
        import export as exp
        assert exp._date_or_none(["2000-01-15"], 0) == "2000-01-15"

    def test_returns_none_for_empty_string(self):
        import export as exp
        assert exp._date_or_none([""], 0) is None

    def test_returns_none_out_of_bounds(self):
        import export as exp
        assert exp._date_or_none([], 0) is None


# ── _resident_row ─────────────────────────────────────────────────────────────

class TestResidentRow:
    def _call(self, resident, addresses=None):
        import export as exp
        exp._ADDRESS_CACHE.clear()
        if addresses is None:
            addresses = [Address(id=resident.address_id, street="Test St 1")]
        mock_db = MagicMock()
        mock_db.get_addresses.return_value = addresses
        with patch.dict("sys.modules", {"database": mock_db}):
            return exp._resident_row(resident)

    def test_active_status_label(self):
        lang.set_lang("en")
        r = _make_resident(status="active")
        row = self._call(r)
        assert row[3] == "active"

    def test_deceased_status_label(self):
        lang.set_lang("en")
        r = _make_resident(status="deceased")
        row = self._call(r)
        assert row[3] == "deceased"

    def test_left_status_label(self):
        lang.set_lang("en")
        r = _make_resident(status="left")
        row = self._call(r)
        assert row[3] == "left"

    def test_ukrainian_status_deceased(self):
        lang.set_lang("uk")
        import export as exp
        exp._ADDRESS_CACHE.clear()
        r = _make_resident(status="deceased")
        mock_db = MagicMock()
        mock_db.get_addresses.return_value = [Address(id=1, street="Test St 1")]
        with patch.dict("sys.modules", {"database": mock_db}):
            row = exp._resident_row(r)
        assert row[3] == "помер"

    def test_date_fields_empty_when_none(self):
        r = _make_resident(birth_date=None, baptism_date=None, marriage_date=None, death_date=None)
        row = self._call(r)
        assert row[4] == ""  # dob
        assert row[5] == ""  # baptism
        assert row[6] == ""  # marriage
        assert row[7] == ""  # death

    def test_date_fields_populated(self):
        r = _make_resident(
            birth_date="1990-01-01", baptism_date="1990-02-01",
            marriage_date="2010-06-15", death_date=None
        )
        row = self._call(r)
        assert row[4] == "1990-01-01"
        assert row[5] == "1990-02-01"
        assert row[6] == "2010-06-15"

    def test_name_fields(self):
        r = _make_resident(first_name="Maria", last_name="Petrenko")
        row = self._call(r)
        assert row[0] == "Petrenko"
        assert row[1] == "Maria"


# ── CSV export ────────────────────────────────────────────────────────────────

class TestExportCsv:
    def _export_and_read(self, tmp_path, residents, addresses=None):
        import export as exp
        exp._ADDRESS_CACHE.clear()
        if addresses is None:
            addresses = [Address(id=1, street="Main St 1")]
        path = str(tmp_path / "out.csv")
        mock_db = MagicMock()
        mock_db.get_addresses.return_value = addresses
        with patch.dict("sys.modules", {"database": mock_db}):
            exp.export_csv(path, residents)
        with open(path, "r", encoding="utf-8") as f:
            return list(csv.reader(f))

    def test_header_row_present(self, tmp_path):
        rows = self._export_and_read(tmp_path, [])
        assert len(rows) == 1  # only header
        assert "Last Name" in rows[0]

    def test_data_row_count(self, tmp_path):
        residents = [_make_resident(id=i) for i in range(3)]
        rows = self._export_and_read(tmp_path, residents)
        assert len(rows) == 4  # 1 header + 3 data

    def test_data_sorted_by_last_then_first(self, tmp_path):
        residents = [
            _make_resident(first_name="Zoriana", last_name="Bila"),
            _make_resident(first_name="Anna", last_name="Bila"),
            _make_resident(first_name="Ivan", last_name="Adamenko"),
        ]
        rows = self._export_and_read(tmp_path, residents)
        last_names = [r[0] for r in rows[1:]]
        assert last_names[0] == "Adamenko"
        assert last_names[1] == "Bila"

    def test_clears_address_cache_before_export(self, tmp_path):
        import export as exp
        exp._ADDRESS_CACHE[9999] = "stale"
        residents = [_make_resident()]
        self._export_and_read(tmp_path, residents)
        assert 9999 not in exp._ADDRESS_CACHE


# ── CSV import via _import_rows ───────────────────────────────────────────────

class TestImportRows:
    def _run(self, db, rows):
        import export as exp
        exp._ADDRESS_CACHE.clear()
        with patch("export.db", db, create=True):
            # _import_rows imports `database as db` internally, so patch at module level
            with patch("database.DB_PATH", db.DB_PATH if hasattr(db, "DB_PATH") else ":memory:"):
                return exp._import_rows(rows)

    def test_skips_empty_rows(self, db):
        import export as exp
        new, skip = exp._import_rows([])
        assert new == 0 and skip == 0

    def test_skips_rows_with_missing_fields(self, db):
        import export as exp
        with patch("database.DB_PATH", db.DB_PATH if hasattr(db, "DB_PATH") else ":memory:"):
            new, skip = exp._import_rows([["", "Ivan", "Main St 1"]])
        assert new == 0

    def test_imports_new_resident(self, db, tmp_path):
        import export as exp
        # patch the db module that _import_rows uses
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            rows = [["Kovalenko", "Ivan", "Main St 1", "active", "1980-01-01", "", "", "", "note"]]
            new, skip = exp._import_rows(rows)
        assert new == 1
        assert skip == 0

    def test_skips_duplicate_resident(self, db, tmp_path):
        import export as exp
        db_path = str(tmp_path / "test2.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            rows = [["Kovalenko", "Ivan", "Main St 1", "active", "", "", "", "", ""]]
            exp._import_rows(rows)  # first import
            new, skip = exp._import_rows(rows)  # second import same data
        assert skip == 1
        assert new == 0

    def test_status_mapping_ukrainian(self, tmp_path):
        import export as exp
        db_path = str(tmp_path / "test3.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            rows = [["Мельник", "Oksana", "Шевченка 3", "активний", "", "", "", "", ""]]
            new, skip = exp._import_rows(rows)
            residents = _db.get_all_residents()
        assert new == 1
        assert residents[0].status == "active"

    def test_status_mapping_deceased(self, tmp_path):
        import export as exp
        db_path = str(tmp_path / "test4.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            rows = [["Smith", "John", "Oak Ave 1", "deceased", "", "", "", "", ""]]
            exp._import_rows(rows)
            residents = _db.get_all_residents()
        assert residents[0].status == "deceased"

    def test_unknown_status_defaults_to_active(self, tmp_path):
        import export as exp
        db_path = str(tmp_path / "test5.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            rows = [["Jones", "Amy", "Pine St 2", "unknown_status", "", "", "", "", ""]]
            exp._import_rows(rows)
            residents = _db.get_all_residents()
        assert residents[0].status == "active"


# ── Full CSV round-trip ───────────────────────────────────────────────────────

class TestCsvRoundTrip:
    def test_export_then_import(self, tmp_path):
        import export as exp
        db_path = str(tmp_path / "roundtrip.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            addr = _db.add_address("Round Trip St 1")
            r = Resident(
                id=None, address_id=addr.id,
                first_name="Taras", last_name="Shevchenko",
                birth_date="1814-03-09", status="active", notes="Poet"
            )
            _db.add_resident(r)

        # Export
        csv_path = str(tmp_path / "export.csv")
        with patch("database.DB_PATH", db_path):
            import database as _db
            residents = _db.get_all_residents()
            addresses = _db.get_addresses()
            exp._ADDRESS_CACHE.clear()
            mock_db = MagicMock()
            mock_db.get_addresses.return_value = addresses
            with patch.dict("sys.modules", {"database": mock_db}):
                exp.export_csv(csv_path, residents)

        # Import into a fresh DB
        fresh_path = str(tmp_path / "fresh.db")
        with patch("database.DB_PATH", fresh_path):
            import database as _db2
            _db2.init_db()
            new, skip = exp.import_csv(csv_path)
            imported = _db2.get_all_residents()

        assert new == 1
        assert skip == 0
        assert imported[0].first_name == "Taras"
        assert imported[0].last_name == "Shevchenko"
        assert imported[0].birth_date == "1814-03-09"


# ── Excel export ──────────────────────────────────────────────────────────────

class TestExportExcel:
    def _export_and_read_ws(self, tmp_path, residents, addresses=None):
        import export as exp
        import openpyxl
        exp._ADDRESS_CACHE.clear()
        if addresses is None:
            addresses = [Address(id=1, street="Main St 1")]
        path = str(tmp_path / "out.xlsx")
        mock_db = MagicMock()
        mock_db.get_addresses.return_value = addresses
        with patch.dict("sys.modules", {"database": mock_db}):
            exp.export_excel(path, residents)
        return openpyxl.load_workbook(path).active, path

    def test_creates_xlsx_file(self, tmp_path):
        import export as exp
        path = str(tmp_path / "out.xlsx")
        mock_db = MagicMock()
        mock_db.get_addresses.return_value = [Address(id=1, street="Main St 1")]
        with patch.dict("sys.modules", {"database": mock_db}):
            exp.export_excel(path, [])
        assert os.path.exists(path)

    def test_header_row_present(self, tmp_path):
        ws, _ = self._export_and_read_ws(tmp_path, [])
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert "Last Name" in headers

    def test_header_is_bold(self, tmp_path):
        ws, _ = self._export_and_read_ws(tmp_path, [])
        assert ws.cell(row=1, column=1).font.bold is True

    def test_data_row_count(self, tmp_path):
        residents = [_make_resident(id=i) for i in range(3)]
        ws, _ = self._export_and_read_ws(tmp_path, residents)
        assert ws.max_row == 4  # 1 header + 3 data rows

    def test_empty_residents_only_header(self, tmp_path):
        ws, _ = self._export_and_read_ws(tmp_path, [])
        assert ws.max_row == 1

    def test_data_sorted_by_last_then_first(self, tmp_path):
        residents = [
            _make_resident(first_name="Zoriana", last_name="Bila"),
            _make_resident(first_name="Anna",    last_name="Bila"),
            _make_resident(first_name="Ivan",    last_name="Adamenko"),
        ]
        ws, _ = self._export_and_read_ws(tmp_path, residents)
        assert ws.cell(row=2, column=1).value == "Adamenko"
        assert ws.cell(row=3, column=1).value == "Bila"

    def test_date_fields_written(self, tmp_path):
        r = _make_resident(
            birth_date="1990-01-01", baptism_date="1990-02-01",
            marriage_date="2010-06-15", death_date=None,
        )
        ws, _ = self._export_and_read_ws(tmp_path, [r])
        assert ws.cell(row=2, column=5).value == "1990-01-01"
        assert ws.cell(row=2, column=6).value == "1990-02-01"
        assert ws.cell(row=2, column=7).value == "2010-06-15"
        # openpyxl stores empty strings as None when read back
        assert ws.cell(row=2, column=8).value in (None, "")

    def test_clears_address_cache_before_export(self, tmp_path):
        import export as exp
        exp._ADDRESS_CACHE[9999] = "stale"
        self._export_and_read_ws(tmp_path, [_make_resident()])
        assert 9999 not in exp._ADDRESS_CACHE

    def test_raises_runtime_error_without_openpyxl(self, tmp_path):
        import export as exp
        path = str(tmp_path / "fail.xlsx")
        with patch.dict("sys.modules", {"openpyxl": None}):
            with pytest.raises(RuntimeError, match="openpyxl"):
                exp.export_excel(path, [])


# ── Excel import ──────────────────────────────────────────────────────────────

class TestImportExcel:
    def _make_xlsx(self, tmp_path, rows, filename="import.xlsx"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Last Name", "First Name", "Address", "Status",
                   "DOB", "Baptism", "Marriage", "Death", "Notes"])
        for row in rows:
            ws.append(row)
        path = str(tmp_path / filename)
        wb.save(path)
        return path

    def test_raises_runtime_error_without_openpyxl(self, tmp_path):
        import export as exp
        path = str(tmp_path / "dummy.xlsx")
        with patch.dict("sys.modules", {"openpyxl": None}):
            with pytest.raises(RuntimeError, match="openpyxl"):
                exp.import_excel(path)

    def test_skips_header_row(self, tmp_path):
        import export as exp
        path = self._make_xlsx(tmp_path, [])
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            new, skip = exp.import_excel(path)
        assert new == 0 and skip == 0

    def test_imports_new_resident(self, tmp_path):
        import export as exp
        rows = [["Kovalenko", "Ivan", "Main St 1", "active", "1980-01-01", "", "", "", "note"]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            new, skip = exp.import_excel(path)
            residents = _db.get_all_residents()
        assert new == 1
        assert skip == 0
        assert residents[0].last_name == "Kovalenko"
        assert residents[0].first_name == "Ivan"
        assert residents[0].notes == "note"

    def test_skips_duplicate_resident(self, tmp_path):
        import export as exp
        rows = [["Petrenko", "Maria", "Oak Ave 2", "active", "", "", "", "", ""]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            exp.import_excel(path)          # first import
            new, skip = exp.import_excel(path)  # second import same data
        assert skip == 1 and new == 0

    def test_date_fields_imported(self, tmp_path):
        import export as exp
        rows = [["Franko", "Ivan", "Franko St 1", "active",
                 "1856-08-27", "1870-01-01", "", "1916-07-28", ""]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            exp.import_excel(path)
            residents = _db.get_all_residents()
        assert residents[0].birth_date == "1856-08-27"
        assert residents[0].baptism_date == "1870-01-01"
        assert residents[0].death_date == "1916-07-28"

    def test_status_mapping_ukrainian(self, tmp_path):
        import export as exp
        rows = [["Мельник", "Oksana", "Шевченка 3", "активний", "", "", "", "", ""]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            exp.import_excel(path)
            residents = _db.get_all_residents()
        assert residents[0].status == "active"

    def test_status_mapping_deceased(self, tmp_path):
        import export as exp
        rows = [["Smith", "John", "Oak Ave 1", "deceased", "", "", "", "", ""]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            exp.import_excel(path)
            residents = _db.get_all_residents()
        assert residents[0].status == "deceased"

    def test_status_mapping_left_ukrainian(self, tmp_path):
        import export as exp
        rows = [["Бондар", "Petro", "Лесі Українки 7", "виїхав", "", "", "", "", ""]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            exp.import_excel(path)
            residents = _db.get_all_residents()
        assert residents[0].status == "left"

    def test_unknown_status_defaults_to_active(self, tmp_path):
        import export as exp
        rows = [["Jones", "Amy", "Pine St 2", "unknown_status", "", "", "", "", ""]]
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            exp.import_excel(path)
            residents = _db.get_all_residents()
        assert residents[0].status == "active"

    def test_skips_rows_with_missing_fields(self, tmp_path):
        import export as exp
        rows = [["", "Ivan", "Main St 1"]]  # missing last name
        path = self._make_xlsx(tmp_path, rows)
        db_path = str(tmp_path / "test.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            new, skip = exp.import_excel(path)
        assert new == 0


# ── Full Excel round-trip ─────────────────────────────────────────────────────

class TestExcelRoundTrip:
    def test_export_then_import(self, tmp_path):
        import export as exp
        db_path = str(tmp_path / "roundtrip.db")
        with patch("database.DB_PATH", db_path):
            import database as _db
            _db.init_db()
            addr = _db.add_address("Round Trip Ave 1")
            r = Resident(
                id=None, address_id=addr.id,
                first_name="Lesia", last_name="Ukrainka",
                birth_date="1871-02-25", baptism_date=None,
                marriage_date=None, death_date="1913-08-01",
                status="deceased", notes="Poet",
            )
            _db.add_resident(r)

        # Export to xlsx
        xlsx_path = str(tmp_path / "export.xlsx")
        with patch("database.DB_PATH", db_path):
            import database as _db
            residents = _db.get_all_residents()
            addresses = _db.get_addresses()
            exp._ADDRESS_CACHE.clear()
            mock_db = MagicMock()
            mock_db.get_addresses.return_value = addresses
            with patch.dict("sys.modules", {"database": mock_db}):
                exp.export_excel(xlsx_path, residents)

        # Import into a fresh DB
        fresh_path = str(tmp_path / "fresh.db")
        with patch("database.DB_PATH", fresh_path):
            import database as _db2
            _db2.init_db()
            new, skip = exp.import_excel(xlsx_path)
            imported = _db2.get_all_residents()

        assert new == 1
        assert skip == 0
        assert imported[0].first_name == "Lesia"
        assert imported[0].last_name == "Ukrainka"
        assert imported[0].birth_date == "1871-02-25"
        assert imported[0].death_date == "1913-08-01"
        assert imported[0].status == "deceased"
